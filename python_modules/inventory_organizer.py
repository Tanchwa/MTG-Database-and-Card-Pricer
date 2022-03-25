import openpyxl
import card_info


mtg_card_inventories = openpyxl.load_workbook("..\Magic Card Spreadsheet.xlsx")
card_list = mtg_card_inventories["mtg_cards"]

for card in range(2, card_list.max_row + 1):
    card_name = card_list.cell(card, 2).value
    set_name = card_list.cell(card, 3).value
    if card_list.cell(card, 4).value == True:
        print(card_name, set_name, "Foil UPDATED")
        card_list.cell(card, 5).value = card_info.card_price_lookup(card_name, set_name, True)
    else:
        print(card_name, set_name, "UPDATED")
        card_list.cell(card, 5).value = card_info.card_price_lookup(card_name,set_name)

mtg_card_inventories.save(filename="..\Magic Card Spreadsheet Appended.xlsx")

